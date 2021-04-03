from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect, request
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View
from apps.habitacion.models import Habitacion, TipoHabitacion
from django.urls import reverse_lazy
from apps.habitacion.forms import HabitacionForm, TipoHabitacionForm
from django.conf import settings
from io import BytesIO
from reportlab .pdfgen import canvas
from reportlab.platypus.tables import Table, TableStyle
from reportlab.platypus import SimpleDocTemplate, Table
from reportlab.lib.units import cm
from reportlab.lib import colors
from django.views.generic.base import TemplateView
from openpyxl import Workbook
# Create your views here.

def index(request):
    return HttpResponse("Index de habitacion")

class TipoHabitacionList(ListView):
    model = TipoHabitacion
    template_name = 'habitacion/tipo/tipo_list.html'

class TipoHabitacionInsert(CreateView):
    model = TipoHabitacion
    template_name = 'habitacion/tipo/tipo_form.html'
    form_class = TipoHabitacionForm
    success_url = reverse_lazy('tipo_listar')

class TipoHabitacionUpdate(UpdateView):
    model = TipoHabitacion
    template_name = 'habitacion/tipo/tipo_form.html'
    form_class = TipoHabitacionForm
    success_url = reverse_lazy('tipo_listar')

class TipoHabitacionDelete(DeleteView):
    model = TipoHabitacion
    template_name = 'habitacion/tipo/tipo_delete.html'
    success_url = reverse_lazy('tipo_listar')

class HabitacionList(ListView):
    model = Habitacion
    template_name = 'habitacion/habitacion_list.html'
    
class HabitacionCreate(CreateView):
    model = Habitacion
    template_name = 'habitacion/habitacion_form.html'
    form_class = HabitacionForm
    success_url = reverse_lazy('habitacion_listar')

class HabitacionUpdate(UpdateView):
    model = Habitacion
    template_name = 'habitacion/habitacion_form.html'
    form_class = HabitacionForm
    success_url = reverse_lazy('habitacion_listar')

class HabitacionDelete(DeleteView):
    model = Habitacion
    template_name = 'habitacion/habitacion_delete.html'
    success_url = reverse_lazy('habitacion_listar')

class ReporteHabitacionPdf(View):
    def cabecera(self, pdf):
        archivo_imagen = settings.MEDIA_ROOT+'\hotelsoft.png'
        pdf.drawImage(archivo_imagen, 40, 750, 120, 90, preserveAspectRatio = True)
        pdf.setFont("Helvetica", 16)
        pdf.drawString(230, 790, u"REPORTE DE HOTELSOFT")
        pdf.setFont("Helvetica", 14)
        pdf.drawString(245, 770, u"REPORTE HABITACIONES")

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type = 'application/pdf')
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)
        self.cabecera(pdf)
        y = 600
        self.tabla(pdf, y)
        pdf.showPage()
        pdf.save()
        pdf = buffer.getvalue()
        buffer.close
        response.write(pdf)
        return response

    def tabla(self, pdf, y):
        encabezados = (
            'Numero',
            'Estado',
            'Costo',
            'Tipo',
        )

        detalles = [(
            Habitacion.numero,
            Habitacion.estado,
            Habitacion.costo,
            Habitacion.Tipo.nombre,
        )
        for Habitacion in Habitacion.objects.all()]

        detalle_solicitud = Table([encabezados]+detalles, colWidths=[3*cm,3*cm,3*cm,3*cm,4*cm,4*cm])
        detalle_solicitud.setStyle(TableStyle(
            [
                ('ALIGN', (0,0), (3,0), 'CENTER'),
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('FONTSIZE', (0,0), (-1,-1), 10),
            ]
        ))

        detalle_solicitud.wrapOn(pdf, 800, 600)
        detalle_solicitud.drawOn(pdf, 150, y)

class ReporteHabitacionExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        from apps.habitacion.models import Habitacion
        detalle = Habitacion.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE HABITACIONES'
        ws.merge_cells('B1:E1')
        ws['B3'] = 'NUMERO'
        ws['C3'] = 'ESTADO'
        ws['D3'] = 'COSTO'
        ws['E3'] = 'TIPO'
        cont = 5

        for Cliente in detalle:
            ws.cell(row = cont, column = 2).value = Habitacion.numero
            ws.cell(row = cont, column = 3).value = Habitacion.estado
            ws.cell(row = cont, column = 4).value = Habitacion.costo
            ws.cell(row = cont, column = 5).value = Habitacion.Tipo.nombre
        cont = cont+1

        nombre_archivo = "ReporteHabitacionesExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    



