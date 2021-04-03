from django.shortcuts import render,redirect
from django.http import HttpResponse, HttpResponseRedirect, request
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View
from openpyxl.workbook.workbook import Workbook
from apps.cliente.models import Cliente, Nacionalidad
from django.urls import reverse_lazy
from apps.cliente.forms import ClienteForm, NacionalidadForm
from django.conf import settings
from io import BytesIO
from reportlab .pdfgen import canvas
from reportlab.platypus.tables import Table, TableStyle
from reportlab.platypus import SimpleDocTemplate, Table
from reportlab.lib.units import cm
from reportlab.lib import colors
from django.views.generic.base import TemplateView
from openpyxl import Workbook


# Create your views here.

def index(request):
    return HttpResponse("Index de cliente")

class ClienteList(ListView):
    model = Cliente
    template_name = 'cliente/cliente_list.html'

class ClienteInsert(CreateView):
    model = Cliente
    template_name = 'cliente/cliente_form.html'
    form_class = ClienteForm
    second_form_class = NacionalidadForm
    success_url = reverse_lazy('cliente_listar')

    def get_context_data(self, **kwargs):
        context = super(ClienteInsert, self).get_context_data()
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(self.request.GET)
        return context

    def post(self, request, *arg, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST)
        form2 = self.second_form_class(request.POST)
        if form.is_valid() and form2.is_valid():
            cliente = form.save(commit = False)
            cliente.fnacionalidad = form2.save()
            cliente.save()
            return HttpResponseRedirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form = form, form2 = form2))

class ClienteUpdate(UpdateView):
    model = Cliente
    second_model = Nacionalidad
    template_name = 'cliente/cliente_form.html'
    form_class = ClienteForm
    second_form_class = NacionalidadForm
    success_url = reverse_lazy('cliente_listar')

    def get_context_data(self, **kwargs):
        context = super(ClienteUpdate, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk', 0)
        cliente = self.model.objects.get(id = pk)
        nacionalidad = self.second_model.objects.get(id = cliente.fnacionalidad_id)
        if 'form' not in context:
            context ['form'] = self.form_class()
        if 'form2' not in context:
            context ['form2'] = self.second_form_class(instance = nacionalidad)
        context['id'] = pk
        return context
    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_cliente = kwargs['pk']
        cliente = self.model.objects.get(id = id_cliente)
        nacionalidad = self.second_model.objects.get(id = cliente.fnacionalidad_id)
        form = self.form_class(request.POST, instance = cliente)
        form2 = self.second_form_class(request.POST, instance = nacionalidad)
        if form.is_valid() and form2.is_valid():
            form.save()
            form2.save()
            return HttpResponseRedirect(self.get_success_url())
        else:
            return HttpResponseRedirect(self.get_success_url())

class ClienteDelete(DeleteView):
    model = Cliente
    template_name = 'cliente/cliente_delete.html'
    form_class = ClienteForm
    second_form_class = NacionalidadForm
    success_url = reverse_lazy('cliente_listar')

class ReporteClientePdf(View):
    def cabecera(self, pdf):
        archivo_imagen = settings.MEDIA_ROOT+'\hotelsoft.png'
        pdf.drawImage(archivo_imagen, 40, 750, 120, 90, preserveAspectRatio = True)
        pdf.setFont("Helvetica", 16)
        pdf.drawString(230, 790, u"REPORTE DE HOTELSOFT")
        pdf.setFont("Helvetica", 14)
        pdf.drawString(260, 770, u"REPORTE CLIENTES")

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type = 'application/pdf')
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)
        self.cabecera(pdf)
        y = 600
        self.tabla(pdf, y)
        pdf.showPage()
        pdf.save()
        pdf = buffer.getvalue()
        buffer.close
        response.write(pdf)
        return response

    def tabla(self, pdf, y):
        encabezados = (
            'Nombre',
            'Direccion',
            'Documento',
            'Telefono',
            'Nacionalidad',
        )

        detalles = [(
            Cliente.nombre,
            Cliente.direccion,
            Cliente.documento,
            Cliente.telefono,
            Cliente.fnacionalidad.nacionalidad,
        )
        for Cliente in Cliente.objects.all()]

        detalle_solicitud = Table([encabezados]+detalles, colWidths=[3*cm,3*cm,3*cm,3*cm,4*cm,4*cm])
        detalle_solicitud.setStyle(TableStyle(
            [
                ('ALIGN', (0,0), (3,0), 'CENTER'),
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('FONTSIZE', (0,0), (-1,-1), 10),
            ]
        ))

        detalle_solicitud.wrapOn(pdf, 800, 600)
        detalle_solicitud.drawOn(pdf, 60, y)

class ReporteClienteExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        from apps.cliente.models import Cliente
        detalle = Cliente.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE CLIENTES'
        ws.merge_cells('B1:F1')
        ws['B3'] = 'NOMBRE'
        ws['C3'] = 'DIRECCION'
        ws['D3'] = 'DOCUMENTO'
        ws['E3'] = 'TELEFONO'
        ws['F3'] = 'NACIONALIDAD'
        cont = 6

        for Cliente in detalle:
            ws.cell(row = cont, column = 2).value = Cliente.nombre
            ws.cell(row = cont, column = 3).value = Cliente.direccion
            ws.cell(row = cont, column = 4).value = Cliente.documento
            ws.cell(row = cont, column = 5).value = Cliente.telefono
            ws.cell(row = cont, column = 6).value = Cliente.fnacionalidad.nacionalidad
        cont = cont+1

        nombre_archivo = "ReporteClientesExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response