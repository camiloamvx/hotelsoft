from django.shortcuts import render
from django.shortcuts import render,redirect
from django.http import HttpResponse, HttpResponseRedirect, request
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View
from apps.alquiler.models import Registrador, Estado, Alquiler
from django.urls import reverse_lazy
from apps.alquiler.forms import RegistradorForm, EstadoForm, AlquilerForm
from apps.cliente.forms import ClienteForm
from apps.habitacion.forms import HabitacionForm
from django.conf import settings
from io import BytesIO
from reportlab .pdfgen import canvas
from reportlab.platypus.tables import Table, TableStyle
from reportlab.platypus import SimpleDocTemplate, Table
from reportlab.lib.units import cm
from reportlab.lib import colors
from django.views.generic.base import TemplateView
from openpyxl import Workbook

# Create your views here.

def index(request):
    return HttpResponse("Index de alquiler")

class RegistradorList(ListView):
    model = Registrador
    template_name = 'alquiler/registrador/registrador_list.html'

class RegistradorInsert(CreateView):
    model = Registrador
    template_name = 'alquiler/registrador/registrador_form.html'
    form_class = RegistradorForm
    success_url = reverse_lazy('registrador_listar')
    
class RegistradorEdit(UpdateView):
    model = Registrador
    template_name = 'alquiler/registrador/registrador_form.html'
    form_class = RegistradorForm
    success_url = reverse_lazy('registrador_listar')

class RegistradorDelete(DeleteView):
    model = Registrador
    template_name = 'alquiler/registrador/registrador_delete.html'
    success_url = reverse_lazy('registrador_listar')

class AlquilerList(ListView):
    model = Alquiler
    template_name = 'alquiler/alquiler_list.html'

class AlquilerInsert(CreateView):
    model = Alquiler
    template_name = 'alquiler/alquiler_form.html'
    form_class = AlquilerForm
    success_url = reverse_lazy('alquiler_listar')
    
class AlquilerUpdate(UpdateView):
    model = Alquiler
    template_name = 'alquiler/alquiler_form.html'
    form_class = AlquilerForm
    success_url = reverse_lazy('alquiler_listar')

class AlquilerDelete(DeleteView):
    model = Alquiler
    template_name = 'alquiler/alquiler_delete.html'
    success_url = reverse_lazy('alquiler_listar')

class ReporteAlquilerPdf(View):
    def cabecera(self, pdf):
        archivo_imagen = settings.MEDIA_ROOT+'\hotelsoft.png'
        pdf.drawImage(archivo_imagen, 40, 750, 120, 90, preserveAspectRatio = True)
        pdf.setFont("Helvetica", 16)
        pdf.drawString(230, 790, u"REPORTE DE HOTELSOFT")
        pdf.setFont("Helvetica", 14)
        pdf.drawString(255, 770, u"REPORTE ALQUILERES")

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type = 'application/pdf')
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)
        self.cabecera(pdf)
        y = 600
        self.tabla(pdf, y)
        pdf.showPage()
        pdf.save()
        pdf = buffer.getvalue()
        buffer.close
        response.write(pdf)
        return response

    def tabla(self, pdf, y):
        encabezados = (
            'Fecha entrada',
            'Fecha salida',
            'Costo total',
            'Habitacion',
            'Cliente',
            'Registrador',
            'Estado',
        )

        detalles = [(
            Alquiler.fechaHoraEntrada,
            Alquiler.fechaHoraSalida,
            Alquiler.costoTotal,
            Alquiler.habitacion.Tipo,
            Alquiler.cliente.nombre,
            Alquiler.registrador.nombre,
            Alquiler.estado.nombre,
        )
        for Alquiler in Alquiler.objects.all()]

        detalle_solicitud = Table([encabezados]+detalles, colWidths=[3*cm,3*cm,3*cm,2*cm,2*cm,3*cm,2*cm])
        detalle_solicitud.setStyle(TableStyle(
            [
                ('ALIGN', (0,0), (3,0), 'CENTER'),
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('FONTSIZE', (0,0), (-1,-1), 10),
            ]
        ))

        detalle_solicitud.wrapOn(pdf, 800, 600)
        detalle_solicitud.drawOn(pdf, 50, y)

class ReporteAlquilerExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        from apps.alquiler.models import Alquiler
        detalle = Alquiler.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE ALQUILERES'
        ws.merge_cells('B1:H1')
        ws['B3'] = 'FECHA ENTRADA'
        ws['C3'] = 'FECHA SALIDA'
        ws['D3'] = 'COSTO TOTAL'
        ws['E3'] = 'HABITACION'
        ws['F3'] = 'CLIENTE'
        ws['G3'] = 'REGISTRADOR'
        ws['H3'] = 'ESTADO'
        cont = 8

        for Alquiler in detalle:
            ws.cell(row = cont, column = 2).value = Alquiler.fechaHoraEntrada
            ws.cell(row = cont, column = 3).value = Alquiler.fechaHoraSalida
            ws.cell(row = cont, column = 4).value = Alquiler.costoTotal
            ws.cell(row = cont, column = 5).value = Alquiler.habitacion.Tipo
            ws.cell(row = cont, column = 6).value = Alquiler.cliente.nombre
            ws.cell(row = cont, column = 7).value = Alquiler.registrador.nombre
            ws.cell(row = cont, column = 8).value = Alquiler.estado.nombre
        cont = cont+1

        nombre_archivo = "ReporteAlquileresExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
